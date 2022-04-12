''' 
This module contains functions to help format and plot Arrhenius data.
The data is electrohchemical impedance spectroscopy (EIS) data obtained by a Gamry potentiostat. 
The files are .DTA files and this module plots EIS and IV curves as well as fits and plots DRT using
the Bayes-DRT package
# C-Meisel
'''

'Imports'
import os #operating system useful for navigating directories
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import scipy as scipy
import seaborn as sns

from bayes_drt.inversion import Inverter
from bayes_drt import plotting as bp 

from .fit_drt import map_drt_save,plot_peiss

def arrhenius_plots(folder_loc:str, jar:str, temps:list, area,plot_eis:bool = True, plot_drt:bool = True,
                    drt_peaks:bool = True, thickness:float = 0, rp_plt_type:str = 'ln', which:str = 'core',
                    init_from_ridge:bool = True, re_fit:bool = False, legend_loc:str = 'outside'):
    '''
    Searches though the folder_loc and separates out the EIS files to be used to plot the arrhenius data.
    The EIS files are matched to their corresponding temperature input from temps.
    The EIS files are map-fit using the Bayes-DRT packange and the ohmic and rp values are extracted and used
    to make an Arrhneius plot for the ohmic and polarizatio ressitance. Each curve is linearly fit and the 
    activation energy of is calculated and shown on the plot. The data is then saved in the cell data excel
    file if it does not already exist. If the data does exist then that data will be used (unless re-fit = True).

    This function gives the option to plot the EIS, DRT, and DRT peaks for the Arrhenius EIS files.

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The location of the directory containing the EIS files. 
    jar, str: (path to a directory)
        The location of the directory to contain the DRT map-fits.
    temps, list: 
        List of temperatures that the EIS spectra were taken at. this is input by the user 
        and must be in the same order that the EIS spectra were taken at.
    area, float:
        The active cell area of the cell in cm^2
    plot_eis, bool:
        If true all the EIS spectra used in making the Arrhenius plot will be plotted.
    plot_drt, bool:
        If true the DRT map-fits will be plotted.
    drt_peaks, bool:
        If true all the DRT spectra will be fit using the Bayes-DRT package and the
        resistance of each peak will be plotted.
    thickness, float: (Default is 0)
        The thickness of the electrolyte of the cell in cm. If there is no SEM data on the electrolyte
        thickness then leave the value at it's default value of 0.
    rp_plt_type, str: (Default is 'ln')
        Type of arrhenius plot to make for the polarization resistance plot. The two values are ln or asr.
        The ln plot is a log-log plot and the asr plot is a semilogx plot. The activation energy (Ea) is 
        only calculated if an ln plot is made.
    which, str: (Default is 'core')
        Which data to store. 'core' or 'sample'. Core file sizes are smaller
    init_from_ridge, bool: optional (default: False)
        If True, use the hyperparametric ridge solution to initialize the Bayesian fit.
        Only valid for single-distribution fits
    re-fit, bool: optional (default: False)
        If True, the EIS data will have the DRT fits re-fit and re-stored in the cell data excel file.
    legend_loc, str: optional (default: 'outside')
        The location of the legend. Outside placed the legend outside the figure.
        The other option is 'best' and this uses 'best' to make the best decision on where to place
        the legend insisde the figure.

    Return --> None but 2-5 plots are crated and shown, the EIS data gets fit and saved, and the 
    data to make the Arrhenius plots is saved in the cell data excel file (if not already there)
    '''

    ' --- Finding correct files and formatting --- '
    ahp_eis = [file for file in os.listdir(folder_loc) if file.endswith('.DTA') and file.find('_Ahp__#')!=-1 and file.find('PEIS')!=-1] #Makes a list of all ahp
    cell_name = os.path.basename(folder_loc).split("_", 3)[2]
    ahp_eis = sorted(ahp_eis, key=lambda x: int((x[x.find('Ahp__#')+len('Ahp__#'):x.rfind('.DTA')]))) #Sorts numerically by eis number (temperature)

    ' --- Conducting a DRT Map fit of the files if not already done so, and saving file names --- '
    ahp_map_fits = [] # List to save fit names for later
    pickel_jar = os.listdir(jar)
    pickel_name = 0 # for determining if the file has been fit or not.

    for c,ahp in enumerate(ahp_eis): # For loop to do a map fit on all Arrhenius EIS
        temp = str(temps[c])
        fit_name = cell_name + '_map_fit_Ahp' + temp + 'C.pkl'
        ahp_map_fits.append(fit_name)
        for pickel in pickel_jar: # checks to see if this has already been fit, if so name gets set to 1
            if fit_name == pickel and re_fit == False:
                pickel_name = pickel_name + 1
                break
        if pickel_name == 0:
            map_drt_save(os.path.join(folder_loc,ahp),jar,fit_name,which=which,init_from_ridge=init_from_ridge)

    ' --- Gathering information to make the Arrhenius Plots and saving it in an excel file sheet for this cell --- '
    # --- Checking to see if this has already been done before
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(folder_loc,excel_name)
    sheet_name = 'Arrhenius data'
    exists = False

    if os.path.exists(excel_file)==True: # Looing for the data excel file in the button cell folder
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        wb = load_workbook(excel_file, read_only=True) # Looking for the Arrhenius Data Sheet
        if sheet_name in wb.sheetnames:
            exists = True
    elif os.path.exists(excel_file)==False:
        writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

    if exists == False: # Make the excel data list
        # --- Initializing lists
        ohmic = [] # Ω
        rp = [] # Ω
        tk_1000 = [] #(1/k) 1000 over temperature in K.  From 625-500C (same order as above lists)

        for fit in ahp_map_fits: # Loading inverter files to extract information from the EIS
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar,fit))
                ohmic.append(inv.R_inf)
                rp.append(inv.predict_Rp())
                temp = fit[fit.find('_Ahp')+len('_Ahp'):fit.rfind('C.pkl')]
                tk_1000.append(1000/(int(temp)+273))

        # --- Calculating Ohmic and Polarization area specific resistance
        rp_asr = np.array(rp) * area #ohms*cm^2
        ohmic_asr = np.array(ohmic) * area #ohms*cm^2

        # --- calculating conductivity and the arrhenius plots for the ohmic and polarization resistance
        conductivity = thickness/(ohmic_asr) #This is the conductivity of the electroltye
        ah_cond = np.log(conductivity*(273+np.array(temps))) #The temperature for this step is kelvin, thus 273 is added to the temperatue in celsius
        ah_rp = np.log(rp_asr/(273+np.array(temps)))
        ah_ohmic_asr = np.log(ohmic_asr/(273+np.array(temps)))

        # --- Making a table with all relavent information
        df_table = pd.DataFrame(list(zip(temps, ohmic, conductivity, ohmic_asr, ah_ohmic_asr,
            ah_cond, rp_asr, ah_rp , tk_1000)), 
            columns =['Temperature (C)','Ohmic Resistance (ohm)', 'Conductivity (S/cm)', 'Ohmic ASR (ohm*cm^2)', 'ln(ohmic*cm^2/T)',
            'ln(sigma*T) (SK/cm)', 'Polarization Resistance ASR (ohm*cm^2)', 'ln(ohm*cm^2/T)','tk_1000 (1000/k)']) #\u03C3

        df_table.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this dataframe to a specific worksheet
        writer.save() # Close the Pandas Excel writer and output the Excel file.
        # df_table.to_excel(excel_file,sheet_name=sheet_name,index=False)

    elif exists == True: # Importing the dataframe with the lists if it does already exist and initializing lists
        df = pd.read_excel(excel_file,sheet_name)
        # --- initializing lists
        tk_1000 = df['tk_1000 (1000/k)'].values
        ah_cond = df['ln(sigma*T) (SK/cm)'].values
        ah_ohmic_asr = df['ln(ohmic*cm^2/T)'].values
        rp_asr = df['Polarization Resistance ASR (ohm*cm^2)'].values
        ah_rp = df['ln(ohm*cm^2/T)'].values

    ' --- Making the Ohmic Resistance Arrhenius Plot --- '
    if thickness > 0:
        fig_ohmic = plt.figure()
        plt.rc('font', size=14)
        ax1 = fig_ohmic.add_subplot(111)
        ax2 = ax1.twiny()
        new_tick_locations = tk_1000.values
        ax1.plot(tk_1000,ah_cond,'ko')
        #Aligning the top axes tick marks with the bottom and converting to celcius
        def tick_function(X):
            V = (1000/np.array(X))-273
            return ["%.0f" % z for z in V]
        ax2.set_xticks(new_tick_locations)
        ax2.set_xticklabels(tick_function(new_tick_locations))
        #linear Fit:
        m, b, r, p_value, std_err = scipy.stats.linregress(tk_1000, ah_cond)
        plt.plot(tk_1000, m*tk_1000+b,'r')
        #creating and formatting table:
        row_labels = ['Intercept','Slope','r squared']
        table_values = [[round(b,3)],[round(m,3)],[round(r**2,3)]]
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower center',rowColours= ['lightblue','lightblue','lightblue'])
        table.scale(1,1.6)
        #Axis labels:
        ax1.set_xlabel('1000/T (1/K)')
        ax2.set_xlabel('Temperature (\u00B0C)')
        ax1.set_ylabel('ln(\u03C3*T) (s*K/cm)')
        #Calculating and printing activation energy
        k = 8.617*10**-5 #boltzmanns constant in Ev/K
        Eact = round(m*k*(-1000),3) # this gives the activation energy in eV
        Eacts = f'{Eact}'
        fig_ohmic.text(0.67,0.37,r'$E_a$ ='+Eacts+'eV')
        plt.tight_layout()

    elif thickness == 0:
        fig_ohmic = plt.figure()
        plt.rc('font', size=14)
        ax1 = fig_ohmic.add_subplot(111)
        ax2 = ax1.twiny()
        ax1.plot(tk_1000,ah_ohmic_asr,'ko')
        #Aligning the top axes tick marks with the bottom and converting to celcius
        def tick_function(X):
            V = (1000/X)-273
            return ["%.0f" % z for z in V]
        ax2.set_xticks(tk_1000)
        ax2.set_xticklabels(tick_function(tk_1000))
        #linear Fit:
        m, b, r, p_value, std_err = scipy.stats.linregress(tk_1000, ah_ohmic_asr)
        plt.plot(tk_1000, m*tk_1000+b,'r')
        #creating and formatting table:
        row_labels = ['Intercept','Slope','r squared']
        table_values = [[round(b,3)],[round(m,3)],[round(r**2,3)]]
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',rowColours= ['lightblue','lightblue','lightblue'])
        table.scale(1,1.6)
        #Axis labels:
        ax1.set_xlabel('1000/T (1/K)')
        ax2.set_xlabel('Temperature (\u00B0C)')
        ax1.set_ylabel('ln(\u03A9$_\mathrm{ohmic}*$cm$^2$/T) (\u03A9*cm$^2$/K)')
        #Calculating and printing activation energy
        k = 8.617*10**-5 #boltzmanns constant in Ev/K
        Eact = round(m*k*(1000),3) # this gives the activation energy in eV
        Eacts = f'{Eact}'
        fig_ohmic.text(0.65,0.33,r'$E_a$ ='+Eacts+'eV')
        plt.tight_layout()

    ' --- Making the Polarizatation Resistance Arrhenius Plot --- '
    if rp_plt_type == 'asr':
        x = tk_1000
        y = rp_asr
        fig = plt.figure()
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twiny() #creates a new x axis that is linked to the first y axis
        new_tick_locs = x 
        #plt.yscale('log')
        ax1.semilogy(x,y,'ko')
        def tick_function(X):
            V = (1000/X)-273
            return ["%.0f" % z for z in V]
        ax2.set_xticks(new_tick_locs)
        ax2.set_xticklabels(tick_function(new_tick_locs))
        #linear Fit:
        m, b, r, p_value, std_err = scipy.stats.linregress(x, np.log10(y))
        plt.plot(x, 10**(m*x+b),'r')
        #creating table:
        row_labels = ['Intercept','Slope','r squared']
        table_values = [[round(b,3)],[round(m,3)],[round(r**2,3)]]
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',rowColours= ['gold','gold','gold'])
        table.scale(1,1.6)
        #Axis labels:
        ax1.set_xlabel('1000/T (1/K)')
        ax2.set_xlabel('Temperature (\u00B0C)')
        ax1.set_ylabel('Rp ASR(\u03A9*$cm^2$)')
        plt.tight_layout()

    elif rp_plt_type == 'ln':
        x = tk_1000
        y = ah_rp
        fig = plt.figure()
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twiny() #creates a new x axis that is linked to the first y axis
        new_tick_locs = x #np.array([1.114,1.145,1.179,1.22,1.253,1.294])
        ax1.plot(x,y,'ko')
        def tick_function(X):
            V = (1000/X)-273
            return ["%.0f" % z for z in V]
        ax2.set_xticks(new_tick_locs)
        ax2.set_xticklabels(tick_function(new_tick_locs))
        #linear Fit:
        m, b, r, p_value, std_err = scipy.stats.linregress(x, y)
        plt.plot(x, (m*x+b),'r')
        #creating table:
        row_labels = ['Intercept','Slope','r squared']
        table_values = [[round(b,3)],[round(m,3)],[round(r**2,3)]]
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',rowColours= ['gold','gold','gold'])
        table.scale(1,1.6)
        #Axis labels:
        ax1.set_xlabel('1000/T (1/K)')
        ax2.set_xlabel('Temperature (\u00B0C)')
        ax1.set_ylabel('ln(R$_p$*cm$^2$/T) (\u03A9*cm$^2$/K)')
        #Calculating and printing activation energy
        k = 8.617*10**-5 #boltzmanns constant in Ev/K
        Eact = round(m*k*(1000),3) # this gives the activation energy in eV
        Eacts = f'{Eact}'
        fig.text(0.65,0.33,r'$E_a$ ='+Eacts+'eV')
        plt.tight_layout()

    plt.show()

    ' --- Plotting EIS ---- '
    if plot_eis == True:
        # --- Setting up the color map
        cmap = plt.cm.get_cmap('coolwarm') #cmr.redshift 
        color_space = np.linspace(0,1,len(ahp_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # indicie of the color array

        # --- Plotting
        for eis in reversed(ahp_eis):
            label = str(temps[len(ahp_eis)-c-1])+'C'
            color = cmap(color_space[c])
            plot_peiss(area,label,os.path.join(folder_loc,eis),color=color,legend_loc=legend_loc)
            c = c+1
        plt.show()

    ' --- Plotting DRT --- '
    if plot_drt == True:
        # --- Setting up the color map
        cmap = plt.cm.get_cmap('coolwarm') #cmr.redshift 
        color_space = np.linspace(0,1,len(ahp_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # indicie of the color array

        # --- Plotting
        fig, ax = plt.subplots() #initializing plots for DRT
        for fit in reversed(ahp_map_fits): # Plotting all AHP map fits
            label = str(temps[len(ahp_eis)-c-1])+'C'
            inv = Inverter()
            inv.load_fit_data(os.path.join(jar,fit))
            color = cmap(color_space[c])
            bp.plot_distribution(None,inv,ax,unit_scale='',label = label,color=color)
            c = c + 1

        ax.legend()
        plt.show()

    ' --- DRT peak fitting and plotting --- '
    if drt_peaks == True:
        # --- Checking to see if the peaks have already been fit:
        peak_data = False
        peak_data_sheet = 'Ahp DRT peak fits'
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') # re-writes the writer to ensure openpyxl is the engine so it appends the data

        if os.path.exists(excel_file)==True: # Looing for the data excel file in the button cell folder
            if peak_data_sheet in wb.sheetnames:
                print(wb.sheetnames)
                peak_data = True

        if peak_data == False: # Make the excel data list
            # --- Fitting peaks and appending to a dataframe
            df_tau_r = pd.DataFrame(columns = ['Temperature','Tau','Resistance']) #Initalizing datframe to save temperature
            temp = 0 #Indice of temperature
            for fit in ahp_map_fits: #Loading DRT, fitting peaks, and saving to a dataframe
                # creating inverter and calling fits
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar,fit))
                inv.fit_peaks(prom_rthresh=0.05) # fit the peaks

                # --- obtain time constants from inverters
                info = inv.extract_peak_info()
                tau = inv.extract_peak_info().get('tau_0') # τ/s
                r = inv.extract_peak_info().get('R')*area # Ω*cm2

                i = 0
                for τ in tau:
                    df_tau_r.loc[len(df_tau_r.index)] = [temps[temp], τ, r[i]]
                    i = i+1

                temp = temp+1

            df_tau_r.to_excel(writer, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

        elif peak_data == True: #load the data into a dataframe
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        # ----- plotting
        palette = sns.color_palette("coolwarm", as_cmap=True)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='Temperature',palette = palette,s=69)

        # -- Astetic stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9*cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')
        plt.tight_layout()
        plt.show()

def arrhenius_iv_curves(folder_loc:str, temps:list):
    '''
    Searches through the folder_loc for hte IV curves taken during arrhenius testing. 
    It plots the iv curve and its corresponding power density curve for each temperature

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The location of the directory containing the EIS files. 
    jar, str: (path to a directory)
        The location of the directory to contain the DRT map-fits.
    temps, list: 
        List of temperatures that the EIS spectra were taken at. this is input by the user 
        and must be in the same order that the EIS spectra were taken at.

    Returns --> Nothin, but plots and shows the IV curves and power density curves for each temperature
    '''
    
    # --- Finding correct files and sorting
    ahp_iv = [file for file in os.listdir(folder_loc) if file.endswith('.DTA') and file.find('Ahp')!=-1 and file.find('IV')!=-1] #Makes a list of all ahp
    ahp_iv = sorted(ahp_iv, key=lambda x: int((x[x.find('Ahp_#')+len('Ahp_#'):x.rfind('.DTA')]))) #Sorts numerically by bias
    ahp_iv_loc = []
    for file in ahp_iv:
        ahp_iv_loc.append(os.path.join(folder_loc,file))

    area_list = [area]*len(ahp_iv)

    # --- Adding the degree symbol to the curves
    str_temps = []
    for temp in temps:
        temp = str(temp)
        temp += '\u00B0C'
        str_temps.append(temp)

    # --- Merging curves and conditions
    curves_conditions = tuple(zip(area_list,str_temps,ahp_iv_loc)) 
    cmap = plt.cm.get_cmap('coolwarm_r')
    plot_ivfcs(curves_conditions,print_Wmax=True,cmap=cmap)


